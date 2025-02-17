import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def generate_report(signals, filename):
    # 创建带格式的Excel文件
    wb = Workbook()
    ws = wb.active
    
    # 设置表头格式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    
    # 写入数据
    for r, row in enumerate(pd.DataFrame(signals).itertuples(), 1):
        for c, value in enumerate(row[1:], 1):
            ws.cell(row=r, column=c, value=value)
            
            # 设置买入/卖出颜色标注
            if c == 6:  # action列
                fill_color = '00FF00' if value == 'BUY' else 'FF0000'
                ws.cell(r, c).fill = PatternFill(start_color=fill_color, fill_type='solid')
    
    # 自动调整列宽
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
    
    wb.save(filename)

def generate_market_analysis_report(market_analysis, selected_stocks):
    """生成市场分析报告"""
    try:
        # 准备报告内容
        report_content = {
            "market_trends": market_analysis.get('market_trend', {}),
            "capital_flows": market_analysis.get('capital_flow', {}),
            "sector_performance": market_analysis.get('sector_performance', {}),
            "selected_stocks": selected_stocks.to_dict('records') if not selected_stocks.empty else []
        }
        
        # 生成报告文本
        report_text = "市场分析报告\n"
        report_text += "=" * 50 + "\n\n"
        
        # 1. 市场趋势分析
        report_text += "1. 市场趋势分析\n"
        report_text += "-" * 30 + "\n"
        for index_name, trend_data in report_content['market_trends'].items():
            report_text += f"{index_name}:\n"
            report_text += f"  - 趋势: {trend_data.get('trend', '未知')}\n"
            report_text += f"  - 涨跌幅: {trend_data.get('price_change', 0):.2f}%\n"
            report_text += f"  - 成交量变化: {trend_data.get('volume_change', 0):.2f}%\n\n"
        
        # 2. 资金流向分析
        report_text += "2. 资金流向分析\n"
        report_text += "-" * 30 + "\n"
        north_money = report_content['capital_flows'].get('north_money_net', 0)
        report_text += f"北向资金净流入: {north_money:.2f}亿元\n\n"
        
        # 3. 行业表现分析
        report_text += "3. 行业表现分析\n"
        report_text += "-" * 30 + "\n"
        sector_perf = report_content['sector_performance']
        if sector_perf:
            report_text += "表现最好的行业:\n"
            for sector in sector_perf.get('top_sectors', []):
                report_text += f"  - {sector.get('行业名称', '未知')}: {sector.get('涨跌幅', 0):.2f}%\n"
            
            report_text += "\n表现最差的行业:\n"
            for sector in sector_perf.get('bottom_sectors', []):
                report_text += f"  - {sector.get('行业名称', '未知')}: {sector.get('涨跌幅', 0):.2f}%\n"
        
        # 4. 选股结果
        report_text += "\n4. 选股结果\n"
        report_text += "-" * 30 + "\n"
        if report_content['selected_stocks']:
            for stock in report_content['selected_stocks']:
                report_text += f"股票代码: {stock.get('股票代码', '未知')}\n"
                report_text += f"股票名称: {stock.get('股票名称', '未知')}\n"
                report_text += f"收盘价: {stock.get('收盘价', 0):.2f}\n"
                report_text += f"涨跌幅: {stock.get('涨跌幅', 0):.2f}%\n"
                report_text += f"换手率: {stock.get('换手率', 0):.2f}%\n"
                report_text += "-" * 20 + "\n"
        else:
            report_text += "今日没有符合条件的股票\n"
        
        # 保存报告
        current_date = datetime.now().strftime('%Y%m%d')
        report_file = f"report_{current_date}.txt"
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report_text)
        
        print(f"\n分析报告已生成: {report_file}")
        return True
        
    except Exception as e:
        print(f"\n生成市场分析报告时出错: {e}")
        return False
